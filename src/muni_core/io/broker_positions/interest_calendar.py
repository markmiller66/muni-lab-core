# interest_calendar.py
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import pandas as pd
import numpy as np

def _as_decimal_coupon(x) -> float | None:
    """Accepts 0.04, 4.0, '4.000%', blank. Returns decimal coupon like 0.04.
       Keeps 0 as 0.0 for true zero-coupon bonds.
    """
    if x is None or (isinstance(x, float) and np.isnan(x)):
        return None

    s = str(x).strip()
    if s == "" or s.lower() in {"nan", "none"}:
        return None

    s = s.replace("%", "")
    try:
        v = float(s)
    except Exception:
        return None

    # ✅ keep true zero-coupon
    if v == 0:
        return 0.0

    # percent-points -> decimal
    if v > 1.5:
        v = v / 100.0

    # sanity bounds (allow tiny, but not absurd)
    if v < 0 or v > 0.20:
        return None

    return v


def _infer_par_amount_row(row: pd.Series) -> float | None:
    """
    Try to get PAR/Face consistently:
      1) FACE VALUE (best if provided)
      2) infer from MRKT VALUE / (MRKT PRICE/100)
      3) fallback to QTY
    """
    # 1) FACE VALUE
    for col in ["FACE VALUE", "PAR", "PAR VALUE"]:
        if col in row.index:
            fv = pd.to_numeric(str(row[col]).replace(",", ""), errors="coerce")
            if pd.notna(fv) and fv > 0:
                return float(fv)

    # 2) Infer from market value and price-per-100
    mv = None
    mp = None
    if "MRKT VALUE" in row.index:
        mv = pd.to_numeric(str(row["MRKT VALUE"]).replace(",", ""), errors="coerce")
    if "MRKT PRICE" in row.index:
        mp = pd.to_numeric(str(row["MRKT PRICE"]).replace(",", ""), errors="coerce")

    if pd.notna(mv) and pd.notna(mp) and mv > 0 and mp > 0:
        par_est = float(mv) / (float(mp) / 100.0)
        # Sanity: ignore absurd results
        if par_est > 0 and par_est < 1e9:
            return par_est

    # 3) Fallback to QTY
    if "QTY" in row.index:
        q = pd.to_numeric(str(row["QTY"]).replace(",", ""), errors="coerce")
        if pd.notna(q) and q > 0:
            return float(q)

    return None


MONTHS = ["jan","feb","mar","apr","may","jun","july","aug","sept","oct","nov","dec"]
BROKERS = ["CHASE","FID","MLWM","RJ","SCHWAB","WF"]

def _month_name(ts: pd.Timestamp) -> str:
    # Use month number so labels match MONTHS exactly
    mapping = {
        1: "jan", 2: "feb", 3: "mar", 4: "apr",
        5: "may", 6: "jun", 7: "july", 8: "aug",
        9: "sept", 10: "oct", 11: "nov", 12: "dec",
    }
    return mapping.get(int(ts.month))
def format_interest_monthly_pivot_table(df: pd.DataFrame) -> pd.DataFrame:
    """
    Supports:
      - Month columns as jan..dec (legacy)
      - Month columns as YYYY-MM (new Pivot B)

    Output:
      MultiIndex columns with 2 header rows:
        sum(YYYY-MM) ...   ""   sum(chase) ...
        broker  cusip  YYYY-MM totals ...   CHASE ...
    """
    out = df.copy()

    # normalize cols to lower for matching, but keep originals
    col_map = {c: str(c).strip() for c in out.columns}
    lower = {c: col_map[c].lower() for c in out.columns}

    def get(name: str):
        name = name.lower()
        for c, lo in lower.items():
            if lo == name:
                return c
        return None

    def is_yyyymm(x: str) -> bool:
        s = str(x).strip()
        return (
            len(s) == 7 and s[4] == "-" and
            s[:4].isdigit() and s[5:7].isdigit()
        )

    broker_col = get("broker") or "BROKER"
    cusip_col  = get("cusip")  or "CUSIP"
    totals_col = get("totals") or "totals"
    annualized_col = get("annualized_total")  # optional, may not exist

    # collect month-like columns
    month_cols = []
    for c in out.columns:
        lo = str(c).strip().lower()
        if lo in MONTHS:
            month_cols.append(c)
        elif is_yyyymm(lo):
            month_cols.append(c)

    # broker total columns (CHASE/FID/...)
    broker_sum_cols = []
    for b in BROKERS:
        c = get(b.lower()) or get(b)
        if c is not None:
            broker_sum_cols.append(c)

    # order: broker, cusip, months (sorted if YYYY-MM), totals, annualized_total (if exists), broker sums
    yyyymm_cols = [c for c in month_cols if is_yyyymm(str(c).strip())]
    mon_cols    = [c for c in month_cols if str(c).strip().lower() in MONTHS]

    yyyymm_cols_sorted = sorted(yyyymm_cols, key=lambda x: str(x))
    mon_cols_sorted = mon_cols  # MONTHS are already controlled by list order if you want; keep as-is

    ordered = [broker_col, cusip_col] + yyyymm_cols_sorted + mon_cols_sorted + [totals_col]
    if annualized_col and annualized_col in out.columns:
        ordered.append(annualized_col)
    ordered += broker_sum_cols

    ordered = [c for c in ordered if c in out.columns]
    out = out[ordered]

    # build 2-row headers
    top, bot = [], []
    for c in out.columns:
        lo = str(c).strip().lower()
        if lo == "broker":
            top.append(""); bot.append("broker")
        elif lo == "cusip":
            top.append(""); bot.append("cusip")
        elif lo in MONTHS or is_yyyymm(lo):
            top.append(f"sum({lo})"); bot.append(lo)
        elif lo in ("totals", "annualized_total"):
            top.append(""); bot.append(lo)
        elif lo in [b.lower() for b in BROKERS]:
            top.append(f"sum({lo})"); bot.append(lo.upper())
        else:
            top.append(""); bot.append(str(c))

    out.columns = pd.MultiIndex.from_arrays([top, bot])
    return out



from muni_core.io.broker_positions.config import ENRICHED_OUT, OUTPUT_DIR
from muni_core.io.broker_positions.utils import normalize_columns, clean_cusip


@dataclass
class InterestCalendarConfig:
    start_month: str | None = None   # "2025-01" or None => auto from min PRICING DATE
    end_month: str | None = None     # "2027-12" or None => auto 24 months forward
    freq_per_year: int = 2           # most munis are semiannual
    out_path: Path = Path(OUTPUT_DIR) / "monthly_interest_calendar.xlsx"


def _to_date(s: pd.Series) -> pd.Series:
    return pd.to_datetime(s, errors="coerce").dt.date


def _to_month_start(d: pd.Series) -> pd.Series:
    dt = pd.to_datetime(d, errors="coerce")
    return dt.dt.to_period("M").dt.to_timestamp()


def _infer_first_coupon_from_maturity(maturity: pd.Timestamp, freq_per_year: int) -> pd.Timestamp | None:
    """
    Fallback: infer coupon dates from maturity date only.
    For semiannual: coupon months = maturity month and month-6, same day.
    """
    if pd.isna(maturity):
        return None
    if freq_per_year != 2:
        # Extend later if you need quarterly, etc.
        return None

    # next coupon date relative to "today" is handled elsewhere;
    # here we just return an "anchor" coupon date pattern.
    # Use maturity month/day as one coupon, and maturity-6mo as the other.
    return maturity  # anchor


def _generate_coupon_dates(
    maturity: pd.Timestamp,
    first_coupon: pd.Timestamp | None,
    start: pd.Timestamp,
    end: pd.Timestamp,
    freq_per_year: int,
) -> list[pd.Timestamp]:
    if pd.isna(maturity):
        return []

    step_months = int(12 / freq_per_year)

    if first_coupon is not None and not pd.isna(first_coupon):
        # Build forward from first_coupon until end or maturity
        dates = []
        d = first_coupon
        while d <= end and d <= maturity:
            if d >= start:
                dates.append(d)
            d = d + pd.DateOffset(months=step_months)
        return dates





    # Fallback: infer pattern from maturity
    anchor = _infer_first_coupon_from_maturity(maturity, freq_per_year)
    if anchor is None:
        return []

    # Determine the two semiannual coupon months by walking backward from maturity
    dates = []
    d = maturity
    while d >= start:
        if d <= end:
            dates.append(d)
        d = d - pd.DateOffset(months=step_months)

    # coupon dates descending; return ascending
    dates = sorted(set(dates))
    return dates


def build_monthly_interest_calendar(cfg: InterestCalendarConfig = InterestCalendarConfig()) -> pd.DataFrame:
    pos = pd.read_excel(ENRICHED_OUT, dtype=str)
    pos = normalize_columns(pos)

    # Required fields
    for c in ["CUSIP", "BROKER", "QTY", "COUPON", "MATURITY"]:
        if c not in pos.columns:
            raise ValueError(f"Missing required column in enriched output: {c}")

    # Optional: if you later add MSRB FIRST COUPON into enriched output, this will pick it up.
    # Accepted names (normalized): "INITIAL INT PMT DATE", "FIRST COUPON DATE"
    first_coupon_col = None
    for cand in ["INITIAL INT PMT DATE", "FIRST COUPON DATE"]:
        if cand in pos.columns:
            first_coupon_col = cand
            break

    # Parse numerics/dates
    # --- PAR / QTY ---
    pos["PAR_AMT"] = pos.apply(_infer_par_amount_row, axis=1)
    qty = pd.to_numeric(pos["PAR_AMT"], errors="coerce")

    # --- COUPON ---
    cpn = pos["COUPON"].apply(_as_decimal_coupon).astype(float)

    # --- MATURITY ---
    mat = pd.to_datetime(pos["MATURITY"], errors="coerce")

    if first_coupon_col:
        first_coupon = pd.to_datetime(pos[first_coupon_col], errors="coerce")
    else:
        first_coupon = pd.Series([pd.NaT] * len(pos))
    valid = qty.notna() & cpn.notna() & mat.notna()
    implied_annual = qty * cpn
    # "bad" means: coupon looks malformed (NaN, too big, too tiny *but non-zero*), or crazy annual interest.
    bad = valid & (
            cpn.isna()
            | (cpn > 0.20)
            | ((cpn != 0) & (cpn < 0.0001))  # allow 0, but not weird tiny non-zero
            | (implied_annual > 1_000_000)
    )

    if bad.any():
        print("\n[Interest Calendar][WARN] Suspicious COUPON/PAR implied annual interest (sample):")
        cols = [c for c in ["BROKER", "CUSIP", "QTY", "PAR_AMT", "COUPON", "MRKT PRICE", "MRKT VALUE", "MATURITY"] if
                c in pos.columns]
        print(pos.loc[bad, cols].head(25).to_string(index=False))

    print("\n[Interest Calendar] positions with usable QTY/COUPON/MATURITY by broker:")
    print(pos.loc[valid, "BROKER"].value_counts().to_string())
    print("\n[Interest Calendar] skipped (missing QTY/COUPON/MATURITY) by broker:")
    print(pos.loc[~valid, "BROKER"].value_counts().to_string())

    # Determine calendar range
    pricing = pd.to_datetime(pos["PRICING DATE"], errors="coerce")
    default_start = pricing.min()
    if pd.isna(default_start):
        default_start = pd.Timestamp.today()

    if cfg.start_month:
        start = pd.Timestamp(cfg.start_month + "-01")
    else:
        start = default_start.to_period("M").to_timestamp()

    # Determine calendar range
    pricing = pd.to_datetime(pos["PRICING DATE"], errors="coerce")
    default_start = pricing.min()
    if pd.isna(default_start):
        default_start = pd.Timestamp.today()

    if cfg.start_month:
        start = pd.Timestamp(cfg.start_month + "-01")
    else:
        start = default_start.to_period("M").to_timestamp()

    # ---- ADD THIS HERE ----
    anchor_year = pd.Timestamp(start).year + 1   # next full calendar year relative to your start
    cfg.end_month = f"{anchor_year + 1}-12"      # ensure we include Dec of the 2nd full year
    # -----------------------

    if cfg.end_month:
        end = pd.Timestamp(cfg.end_month + "-01") + pd.offsets.MonthEnd(0)
    else:
        end = (start + pd.DateOffset(months=24)) + pd.offsets.MonthEnd(0)




    if cfg.end_month:
        end = pd.Timestamp(cfg.end_month + "-01") + pd.offsets.MonthEnd(0)
    else:
        end = (start + pd.DateOffset(months=24)) + pd.offsets.MonthEnd(0)

    # Build long cashflow rows
    rows = []
    for i in range(len(pos)):
        if pd.isna(qty.iloc[i]) or pd.isna(cpn.iloc[i]) or pd.isna(mat.iloc[i]):
            continue

        cusip = clean_cusip(pos.loc[i, "CUSIP"])
        broker = str(pos.loc[i, "BROKER"])
        key = f"{cusip}|{broker}"

        coupon_dates = _generate_coupon_dates(
            maturity=mat.iloc[i],
            first_coupon=first_coupon.iloc[i],
            start=start,
            end=end,
            freq_per_year=cfg.freq_per_year,
        )





        # interest per payment
        pay_amt = float(qty.iloc[i]) * float(cpn.iloc[i]) / cfg.freq_per_year




        for d in coupon_dates:
            rows.append({
                "MONTH": d.to_period("M").to_timestamp(),
                "PAY_DATE": d,
                "CUSIP": cusip,
                "BROKER": broker,
                "KEY": key,
                "COUPON": float(cpn.iloc[i]),
                "QTY": float(qty.iloc[i]),
                "INTEREST": pay_amt,
                "MATURITY": mat.iloc[i],
                "FIRST_COUPON_USED": first_coupon.iloc[i] if first_coupon_col else pd.NaT,
                "SCHEDULE_MODE": "msrb_first_coupon" if (first_coupon_col and not pd.isna(first_coupon.iloc[i])) else "inferred_from_maturity",
            })


    long_df = pd.DataFrame(rows)
    if long_df.empty:
        raise RuntimeError("No coupon cashflows generated (check COUPON/QTY/MATURITY parsing).")




    # -------------------------
    # Pivot A (keep): MONTH rows × KEY cols
    # -------------------------
    monthly = long_df.groupby(["MONTH", "KEY"], as_index=False)["INTEREST"].sum()
    pivot_month_by_key = (
        monthly.pivot(index="MONTH", columns="KEY", values="INTEREST")
        .fillna(0.0)
        .sort_index()
    )
    # -------------------------
    # Pivot B: broker/cusip rows × YYYY-MM columns + totals + broker totals
    # -------------------------

    # Base row set = every (BROKER, CUSIP) in the enriched file
    base = pos[["BROKER", "CUSIP"]].copy()
    base["BROKER"] = base["BROKER"].astype(str).str.upper().str.strip()
    base["CUSIP"] = base["CUSIP"].map(clean_cusip)
    base = base.dropna(subset=["CUSIP"]).drop_duplicates()

    m2 = long_df.copy()
    m2["BROKER"] = m2["BROKER"].astype(str).str.upper().str.strip()
    m2["CUSIP"] = m2["CUSIP"].map(clean_cusip)

    # YEAR_MONTH = YYYY-MM from PAY_DATE (preferred)
    m2["YEAR_MONTH"] = pd.to_datetime(m2["PAY_DATE"], errors="coerce").dt.strftime("%Y-%m")

    by_bc_ym = (
        m2.groupby(["BROKER", "CUSIP", "YEAR_MONTH"], as_index=False)["INTEREST"]
        .sum()
    )

    interest_table = (
        by_bc_ym.pivot_table(
            index=["BROKER", "CUSIP"],
            columns="YEAR_MONTH",
            values="INTEREST",
            aggfunc="sum",
            fill_value=0.0,
        )
        .reset_index()
    )

    # Join onto full base set so every broker/cusip prints
    interest_table = base.merge(interest_table, on=["BROKER", "CUSIP"], how="left")

    # Identify YYYY-MM columns
    ym_cols = [c for c in interest_table.columns if c not in ("BROKER", "CUSIP")]

    # Fill missing month cells with 0
    for c in ym_cols:
        interest_table[c] = pd.to_numeric(interest_table[c], errors="coerce").fillna(0.0)

    # Sort YYYY-MM columns chronologically
    ym_cols_sorted = sorted(ym_cols)

    # Horizon totals (sum across the visible horizon months)
    interest_table["totals"] = interest_table[ym_cols_sorted].sum(axis=1)

    # (Optional) Annualize totals so it compares to ANNUAL_INTEREST cleanly
    horizon_years = max(1e-9, len(ym_cols_sorted) / 12.0)
    interest_table["annualized_total"] = (interest_table["totals"] / horizon_years).round(6)

    # broker totals (scalar repeated for each row)
    for b in BROKERS:
        scalar = float(interest_table.loc[interest_table["BROKER"].eq(b), "totals"].sum())
        interest_table[b] = scalar

    # Final ordering
    interest_table = interest_table[["BROKER", "CUSIP"] + ym_cols_sorted + ["totals", "annualized_total"] + BROKERS]

    # final format to your two-row header layout
    interest_table_fmt = format_interest_monthly_pivot_table(interest_table)

    # Metadata sheet (one row per KEY)
    meta = (
        long_df.sort_values(["KEY", "PAY_DATE"])
        .groupby("KEY", as_index=False)
        .agg({
            "CUSIP": "first",
            "BROKER": "first",
            "COUPON": "first",
            "QTY": "first",
            "MATURITY": "first",
            "FIRST_COUPON_USED": "first",
            "SCHEDULE_MODE": "first",
        })
        .sort_values(["CUSIP", "BROKER"])
    )

    # ============================================================
    # Pivot C: Broker-level monthly + annualized interest summary
    #   Print ONLY the first 2 calendar years in the horizon
    #   Add a TOTAL row across brokers
    # ============================================================

    bm = (
        m2.groupby(["BROKER", "YEAR_MONTH"], as_index=False)["INTEREST"]
        .sum()
    )

    broker_month = (
        bm.pivot(index="BROKER", columns="YEAR_MONTH", values="INTEREST")
        .fillna(0.0)
    )

    # month columns in YYYY-MM form, sorted
    month_cols_all = sorted([c for c in broker_month.columns
                             if isinstance(c, str) and len(c) == 7 and c[4] == "-"])

    # choose anchor year = next full calendar year
    anchor_year = pd.Timestamp.today().year + 1
    # ensure horizon reaches the end of the 2nd full calendar year shown
    cfg.end_month = f"{anchor_year + 1}-12"

    # choose full years to show
    years_keep = [str(anchor_year), str(anchor_year + 1)]



    # keep ONLY Jan..Dec for those years (avoid partial-year skew)
    month_cols_keep = [
        f"{y}-{m:02d}"
        for y in years_keep
        for m in range(1, 13)
        if f"{y}-{m:02d}" in month_cols_all
    ]

    broker_month = broker_month[month_cols_keep]

    # add annualized + per-year totals
    for y in years_keep:
        y_cols = [c for c in month_cols_keep if c.startswith(f"{y}-")]
        if y_cols:
            # annualized for that calendar year
            if len(y_cols) == 12:
                broker_month[f"annualized_{y}"] = broker_month[y_cols].sum(axis=1).round(2)
            else:
                broker_month[f"annualized_{y}"] = np.nan

            # TOTAL for that calendar year (no annualization)
            broker_month[f"total_{y}"] = broker_month[y_cols].sum(axis=1).round(2)

    # overall totals (this year + next year)
    broker_month["totals"] = broker_month[month_cols_keep].sum(axis=1).round(2)

    # reorder columns
    ordered = []
    for y in years_keep:
        ordered.append(f"annualized_{y}")
        ordered += [c for c in month_cols_keep if c.startswith(f"{y}-")]
        ordered.append(f"total_{y}")

    ordered.append("totals")
    broker_month = broker_month[ordered]

    # add TOTAL row across brokers
    total_row = broker_month.sum(axis=0, numeric_only=True)
    total_row.name = "total"
    broker_month = pd.concat([broker_month, total_row.to_frame().T], axis=0)

    from datetime import datetime

    # timestamp like 2026-02-18_14-37-09
    ts = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")

    out_path = cfg.out_path
    out_path = out_path.with_name(f"{out_path.stem}_{ts}{out_path.suffix}")

    # Write output workbook

    out_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(out_path, engine="openpyxl") as xw:
        long_df.to_excel(xw, sheet_name="interest_long", index=False)

        # existing pivot (month rows)
        pivot_month_by_key.to_excel(xw, sheet_name="interest_monthly_pivot")

        # your requested layout + 2-row headers
        from openpyxl.utils import get_column_letter

        ws = xw.book.create_sheet("interest_by_cusip_broker")

        # ---- write header rows manually ----
        top, bot = interest_table_fmt.columns.levels
        top_row = list(interest_table_fmt.columns.get_level_values(0))
        bot_row = list(interest_table_fmt.columns.get_level_values(1))

        ws.append(top_row)
        ws.append(bot_row)

        # merge identical top headers
        col_start = 1
        for i in range(1, len(top_row) + 2):
            if i == len(top_row) + 1 or top_row[i - 1] != top_row[col_start - 1]:
                if top_row[col_start - 1]:
                    ws.merge_cells(
                        start_row=1,
                        start_column=col_start,
                        end_row=1,
                        end_column=i - 1,
                    )
                col_start = i

        # ---- write data rows ----
        for row in interest_table_fmt.itertuples(index=False):
            ws.append(list(row))

        broker_month_out = broker_month.reset_index()  # BROKER becomes a real column

        broker_month_out.to_excel(
            xw,
            sheet_name="broker_interest_summary",
            index=False
        )

        meta.to_excel(xw, sheet_name="positions_meta", index=False)


    print(f"Wrote: {out_path} (months={pivot_month_by_key.shape[0]} cols={pivot_month_by_key.shape[1]})")

    return pivot_month_by_key



def main() -> None:
    build_monthly_interest_calendar()


if __name__ == "__main__":
    main()
